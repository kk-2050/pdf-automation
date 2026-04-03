# =============================================================================
# Program Name : utils.py
# Author       : Kaori Kashiwagi
# Date         : 2026-04-03
# Purpose      : Shared utility functions and constants used across all modules
#                of the PDF Keyword Highlight Tool. Includes:
#                  - Color name to RGB conversion for PDF highlights
#                  - Highlight opacity constant
#                  - Keyword string parsing
#                  - Boolean normalization (TRUE/FALSE from Excel)
#                  - Excel log file creation and appending
# =============================================================================

import os
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook

# ── Excel log file column headers ─────────────────────────────────────────────
# These headers are written to the first row of a new log file
LOG_HEADERS = [
    "Customer",      # Customer name from Excel master
    "File Name",     # PDF filename
    "Keywords",      # Comma-separated list of keywords searched
    "Pages Scanned", # Number of pages actually searched
    "Color",         # Highlight color used
    "Begin_Page",    # First page searched
    "End_Page",      # Last page searched (or "all")
    "Moved To",      # Full path of the highlighted PDF in Done folder
    "Status",        # OK_HITS_n / SKIPPED_EXISTS / ERROR / CANCELLED
    "Timestamp",     # Date and time of processing
    "User"           # Windows username who ran the tool
]

# ── Highlight color definitions ───────────────────────────────────────────────
# Colors are defined as RGB tuples with values between 0.0 and 1.0
# These are used by PyMuPDF to set the highlight annotation stroke color
COLOR_MAP = {
    "lightblue": (0.529, 0.808, 0.922),  # Default light blue
    "blue":      (0.529, 0.808, 0.922),  # Same as lightblue
    "yellow":    (1.0,   0.949, 0.0),    # Bright yellow
    "green":     (0.596, 0.984, 0.596),  # Light green
    "pink":      (1.0,   0.714, 0.757),  # Soft pink
    "orange":    (1.0,   0.753, 0.302),  # Orange
    "purple":    (0.800, 0.706, 1.000),  # Light purple
    "red":       (1.0,   0.600, 0.600),  # Soft red
}

# ── Highlight opacity ─────────────────────────────────────────────────────────
# 0.0 = fully transparent, 1.0 = fully opaque
# 0.4 (40%) allows the keyword text to remain readable through the highlight
HIGHLIGHT_OPACITY = 0.4


def color_to_rgb01(color_name: str) -> tuple:
    """
    Convert a color name string to an RGB tuple (values 0.0–1.0).
    Falls back to light blue if the color name is not recognized.

    Example:
        color_to_rgb01("yellow")  →  (1.0, 0.949, 0.0)
        color_to_rgb01("xyz")     →  (0.529, 0.808, 0.922)  # fallback
    """
    return COLOR_MAP.get(color_name.lower().strip(), COLOR_MAP["lightblue"])


def is_whole_word_match(text: str, keyword: str) -> bool:
    """
    Check if a keyword appears as a complete word in the given text.
    Uses word boundary (\\b) regex matching, case-insensitive.

    Example:
        is_whole_word_match("PPAP approved", "PPAP")  →  True
        is_whole_word_match("PPAP-Approved", "PPAP")  →  True
        is_whole_word_match("PPAP2 data", "PPAP")     →  False (part of "PPAP2")
    """
    pattern = r'\b' + re.escape(keyword) + r'\b'
    return bool(re.search(pattern, text, re.IGNORECASE))


def split_keywords(raw: str) -> list:
    """
    Parse a raw keyword string from Excel into a list of individual keywords.
    Handles comma separators, Japanese commas (、), and newlines.
    Strips whitespace and removes empty entries.

    Example:
        split_keywords("PPAP, IMDS, RoHS")  →  ["PPAP", "IMDS", "RoHS"]
        split_keywords("PPAP、IMDS")        →  ["PPAP", "IMDS"]
    """
    if not raw:
        return []
    # Replace Japanese comma and newlines with standard comma
    raw = raw.replace("\u3001", ",").replace("\n", ",")
    return [k.strip() for k in raw.split(",") if k.strip()]


def normalize_bool(val) -> bool:
    """
    Convert an Excel cell value to a Python boolean.
    Accepts TRUE, FALSE, YES, NO, 1, 0 (case-insensitive).
    Returns False for None or unrecognized values.

    Example:
        normalize_bool("TRUE")   →  True
        normalize_bool("FALSE")  →  False
        normalize_bool(None)     →  False
        normalize_bool("yes")    →  True
    """
    if val is None:
        return False
    return str(val).strip().upper() in ("TRUE", "YES", "1")


def _ensure_workbook(log_path: str):
    """
    Open the existing Excel log file, or create a new one with headers
    if it doesn't exist yet.

    This function is used internally by append_log().
    The log file is never deleted — new rows are always appended.
    """
    dir_path = os.path.dirname(log_path)
    if dir_path:
        os.makedirs(dir_path, exist_ok=True)  # Create directory if needed

    if os.path.exists(log_path):
        # Log file already exists — open and return it
        wb = load_workbook(log_path)
        ws = wb.active
        # Add headers if the file is somehow empty
        if ws.max_row < 1:
            ws.append(LOG_HEADERS)
        return wb

    # Log file does not exist — create a new workbook with headers
    wb = Workbook()
    ws = wb.active
    ws.title = "Log"
    ws.append(LOG_HEADERS)
    return wb


def append_log(log_path: str, row: list):
    """
    Append one result row to the shared Excel log file.
    Creates the file with headers if it doesn't exist yet.

    Parameters:
        log_path : Full path to pdf_highlight_log.xlsx
        row      : List of values matching LOG_HEADERS order
    """
    wb = _ensure_workbook(log_path)
    ws = wb.active
    ws.append(row)      # Add the new row at the bottom
    wb.save(log_path)   # Save immediately after each row


def build_log_row(customer, file_name, keywords, pages_scanned,
                  color, begin_page, end_page, moved_to, status, user):
    """
    Build a single log row as a list of values, ready to append to Excel.
    Automatically adds the current timestamp.

    Parameters:
        customer      : Customer name
        file_name     : PDF filename
        keywords      : Comma-separated keyword string
        pages_scanned : Number of pages searched
        color         : Highlight color name
        begin_page    : First page searched
        end_page      : Last page searched (None → "all")
        moved_to      : Destination path of the highlighted PDF
        status        : Processing result (OK_HITS_n / SKIPPED_EXISTS / ERROR)
        user          : Windows username

    Returns:
        List of values in LOG_HEADERS order
    """
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # Current timestamp
    ep = str(end_page) if end_page else "all"           # Show "all" if no end page
    return [customer, file_name, keywords, pages_scanned,
            color, begin_page, ep, moved_to, status, ts, user]
