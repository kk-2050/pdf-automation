# =============================================================================
# Program Name : excel_master.py
# Author       : Kaori Kashiwagi
# Date         : 2026-04-03
# Purpose      : Reads the Excel configuration file (highlight_master_v3.xlsx)
#                and converts each row in the Master sheet into a CustomerRule
#                object used by the processing engine.
#                Handles flexible column name matching (case-insensitive,
#                ignores spaces and underscores) so minor typos don't break it.
# =============================================================================

from dataclasses import dataclass
from typing import List, Optional
from openpyxl import load_workbook
from utils import split_keywords, normalize_bool


@dataclass
class CustomerRule:
    """
    Represents one row from the highlight_master_v3.xlsx Master sheet.
    Each field corresponds to a column in the Excel file.

    Fields:
        customer          : Customer name shown in log (e.g. "CustomerA")
        customer_folder   : Path to the customer folder (e.g. C:\\PDFs\\CustomerA)
        destination_folder: Where highlighted PDFs are moved (e.g. C:\\PDFs\\CustomerA\\Done)
        keywords          : List of keywords to search for (e.g. ["PPAP", "IMDS", "RoHS"])
        case_sensitive    : If True, match exact uppercase/lowercase
        whole_word        : If True, match whole words only
        color             : Highlight color name (e.g. "blue", "yellow")
        begin_page        : First page to search (1-based, default 1)
        end_page          : Last page to search (None = all pages)
    """
    customer: str
    customer_folder: str
    destination_folder: str
    keywords: List[str]
    case_sensitive: bool = False
    whole_word: bool = False
    color: str = "lightblue"
    begin_page: int = 1
    end_page: Optional[int] = None  # None means search all pages


def load_master_rules(xlsx_path: str, sheet_name: str = "Master") -> List[CustomerRule]:
    """
    Read the Excel configuration file and return a list of CustomerRule objects.

    Steps:
      1. Open the Excel file (read-only via data_only=True to get values not formulas)
      2. Read the header row and build a column name → column index map
      3. Loop through data rows and create a CustomerRule for each non-empty row
      4. Return the complete list of rules

    Parameters:
        xlsx_path  : Full path to highlight_master_v3.xlsx
        sheet_name : Name of the sheet to read (default: "Master")

    Returns:
        List of CustomerRule objects, one per data row in the sheet
    """
    wb = load_workbook(xlsx_path, data_only=True)  # data_only=True reads cell values, not formulas

    # Verify the expected sheet exists
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]

    # Build a header map: normalized column name → column index
    # Normalization: lowercase, remove spaces and underscores
    # This allows flexible matching (e.g. "Begin_Page" = "beginpage" = "begin page")
    header = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v:
            key = str(v).strip().lower().replace(" ", "").replace("_", "")
            header[key] = col

    def get(row, key):
        """Get a cell value by normalized column name."""
        k = key.lower().replace("_", "")
        c = header.get(k)
        return ws.cell(row=row, column=c).value if c else None

    def get_int(row, key, default):
        """Get a cell value as an integer, with a fallback default."""
        val = get(row, key)
        try:
            return int(val) if val is not None else default
        except (ValueError, TypeError):
            return default

    # Check that all required columns are present
    required = ["customer", "customerfolder", "destinationfolder", "keywords"]
    missing = [k for k in required if k not in header]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    # Read each data row (starting from row 2, skipping the header)
    rules = []
    for r in range(2, ws.max_row + 1):
        customer = get(r, "customer")

        # Skip empty rows
        if not customer:
            continue

        # Read optional columns with safe fallbacks
        color_val = str(get(r, "color") or "lightblue").strip().lower()
        begin = get_int(r, "beginpage", 1)

        # End_Page: None means "search all pages"
        end_raw = get(r, "endpage")
        end = int(end_raw) if end_raw is not None else None

        # Create a CustomerRule object for this row
        rule = CustomerRule(
            customer=str(customer).strip(),
            customer_folder=str(get(r, "customerfolder") or "").strip(),
            destination_folder=str(get(r, "destinationfolder") or "").strip(),
            keywords=split_keywords(str(get(r, "keywords") or "")),
            case_sensitive=normalize_bool(get(r, "casesensitive")),
            whole_word=normalize_bool(get(r, "wholeword")),
            color=color_val,
            begin_page=max(1, begin),  # Ensure begin_page is at least 1
            end_page=end,
        )
        rules.append(rule)

    return rules
